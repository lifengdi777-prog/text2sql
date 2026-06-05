"""导出渲染器 —— 把「智能助手」的编辑结果**保样式**写回原件副本,产出可下载 xlsx。

思路(见 docs/dataset_edit_agent_design.md §6):不重建 Excel,而是打开**原件副本**,
按 diff(最终态 vs 原始态,duckdb_edit.diff_sheet 产出)用 openpyxl **逐格回写** ——
没动过的格保持原件原样(原始文本/公式/样式),只改/删/加变化的部分。

只处理"数据级编辑 + 加/删/改列",不碰合并单元格/图表/跨 sheet 公式(决策 5)。
"""
from __future__ import annotations

import io
from copy import copy
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from services import object_store
from services.duckdb_edit import EXCEL_ROW, EditWorkbook, diff_sheet


# ───────────────────────── §6.1 镜像原列表示 ─────────────────────────
def _render_like_column(ws: Worksheet, col_idx: int, skip_row: int, value: Any,
                        data_start_row: int) -> Any:
    """看同列某未改单元格的原始形态,把 value 渲染成"长得一样"的写法。

    原列是文本带千分位 → 返回 '14,000' 文本;带 % → '88%' 文本;
    原列是真数字(或空)→ 返回数字本身(沿用该列原 number_format,Excel 右对齐)。
    避免把单格写成真数字而在文本列里右对齐、丢绿三角、显得突兀。
    """
    sample = None
    for r in range(data_start_row, ws.max_row + 1):
        if r == skip_row:
            continue
        v = ws.cell(r, col_idx).value
        if v is not None:
            sample = v
            break

    if isinstance(sample, str) and isinstance(value, (int, float)) and not isinstance(value, bool):
        if sample.strip().endswith("%"):
            num = value if value % 1 else int(value)
            return f"{num}%"
        if "," in sample:
            return f"{int(value):,}" if float(value).is_integer() else f"{value:,}"
        return str(value)
    return value


# ───────────────────────── 单 sheet 回写 ─────────────────────────
def patch_sheet(ws: Worksheet, diff: dict, final_df: pd.DataFrame,
                lineage: dict) -> None:
    """按 diff 把一个 sheet 的变更回写到 ws(保样式)。

    顺序(避免行/列号错位):改值 → 改列名 → 加列 → 删列(倒序)→ 删行(倒序)。
    """
    col_excel: dict[str, int] = (lineage or {}).get("col_excel") or {}
    data_start = (lineage or {}).get("data_start_row", 2)
    header_row = (lineage or {}).get("header_row", 1)

    # ① 改值:定位单元格改 .value,样式天然保留;数值按 §6.1 镜像原列写法
    for ch in diff["cell_changes"]:
        c = col_excel.get(ch["col"])
        er = ch["excel_row"]
        if c is None or er is None:
            continue
        ws.cell(er, c).value = _render_like_column(ws, c, er, ch["new"], data_start)

    # ② 改列名:只改表头格文本,列数据/样式不动
    for rn in diff["renames"]:
        c = rn.get("excel_col")
        if c is not None:
            ws.cell(header_row, c).value = rn["new"]

    # ③ 加列:末列右侧追加,表头 + 每行值,样式克隆左邻列
    new_col_idx: dict[str, int] = {}
    for col_name in diff["added_cols"]:
        idx = ws.max_column + 1
        new_col_idx[col_name] = idx
        left = idx - 1
        h = ws.cell(header_row, idx)
        h.value = col_name
        if left >= 1:
            h._style = copy(ws.cell(header_row, left)._style)
        for _, row in final_df.iterrows():
            er = row[EXCEL_ROW]
            if pd.isna(er):           # 新插入行单独处理(④之外),这里只填有原始行的
                continue
            er = int(er)
            t = ws.cell(er, idx)
            val = row[col_name]
            t.value = None if pd.isna(val) else val
            if left >= 1:
                t._style = copy(ws.cell(er, left)._style)

    # ④ 删列:倒序删(避免左删导致右移),用 diff 提供的原始 Excel 列号
    drop_idx = sorted(
        (col_excel[c] for c in diff["dropped_cols"] if c in col_excel),
        reverse=True,
    )
    for c in drop_idx:
        ws.delete_cols(c, 1)

    # ⑤ 删行:倒序删
    for er in sorted((d["excel_row"] for d in diff["deleted"] if d["excel_row"]),
                     reverse=True):
        ws.delete_rows(er, 1)

    # 注:new_rows(新增整行)暂未在原件物理 insert —— MVP 先支持改/删/加列,
    # 新增行的保样式插入留待后续(需 clone 邻行样式 + 处理 excel_row 偏移)。TODO。


# ───────────────────────── 高层编排 ─────────────────────────
def _find_original_key(folder: str) -> str | None:
    """在 ds_{id}/original/ 下找留档的原件 key(取第一个 .xlsx)。"""
    for key in object_store.list_prefix(f"{folder}/original/"):
        if key.lower().endswith((".xlsx", ".xlsm")):
            return key
    return None


def export_with_info(info: dict, active_ops: list[str]) -> tuple[str, bytes]:
    """重放 active_ops → 算 diff → 在原件副本上保样式回写 → 返回 (文件名, xlsx 字节)。

    info 由调用方(async 侧)用 get_dataset_info() 预先取好传入;本函数纯同步
    (openpyxl/boto3 都是同步),调用方用 asyncio.to_thread 包一层。

    步骤:
      1. 取原件字节;
      2. 物化基线(无 op)= 原始态;物化 + replay(active_ops) = 最终态;
      3. 各 sheet diff;
      4. openpyxl 打开原件副本逐 sheet 回写;
      5. 返回字节。
    """
    folder = info["folder_path"]
    key = _find_original_key(folder)
    if key is None:
        raise FileNotFoundError(f"数据集 {info.get('dataset_id')} 未找到原件(original/)")
    original_bytes = object_store.get_bytes(key)
    filename = key.rsplit("/", 1)[-1]

    base = EditWorkbook.from_dataset(info)
    edited = EditWorkbook.from_dataset(info)
    try:
        edited.replay(active_ops)
        diffs = {
            s: diff_sheet(base.current(s), edited.current(s), edited.lineage.get(s))
            for s in edited.sheets()
        }
        finals = {s: edited.current(s) for s in edited.sheets()}
        lineages = {s: edited.lineage.get(s) or {} for s in edited.sheets()}
    finally:
        base.close()
        edited.close()

    wb = load_workbook(io.BytesIO(original_bytes))
    for sheet, diff in diffs.items():
        if sheet in wb.sheetnames:
            patch_sheet(wb[sheet], diff, finals[sheet], lineages[sheet])
    wb.calculation.fullCalcOnLoad = True  # 留存公式让 Excel 打开时重算

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return filename, buf.getvalue()
