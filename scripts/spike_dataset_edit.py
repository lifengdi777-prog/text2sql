"""Phase 0 spike —— 验证 dataset_edit「智能助手」的两条最不确定的链。

不接会话 / SSE / 对象存储,纯本地脚本,只为证明架构可落地:

  链 ①  清洗 → 物化进 DuckDB(带 row_id 身份 + excel_row 物理来源)
        → 跑 UPDATE / DELETE / ALTER ADD COLUMN
        → 对比"最终态 vs 原始态"得到单元格级 diff
  链 ②  openpyxl 打开原件副本 → 按 diff 逐格回写(保样式)→ 存 result.xlsx
        → 重新读回,验证「没改的格保持原始文本(如 13,478)、改过的格是新值」

简化(spike 专用,真实实现会复用 services/excel_ingest):
  - 表头识别:normal_production.xlsx 表头在第 1 行,直接 header=0,不调 LLM。
  - 类型清洗:内联一个迷你版 _coerce_column(去千分位/%/空格试转数值),够验证即可。
  - 只处理一个 sheet「生产明细」,三种操作各来一刀。

用法:
  python scripts/spike_dataset_edit.py [输入xlsx] [输出xlsx]
默认输入 ../../normal_production.xlsx(主仓库根),输出 ./spike_result.xlsx
"""
from __future__ import annotations

import sys
from copy import copy

import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET = "生产明细"
HEADER_ROW = 1          # 表头所在 Excel 行(1-based)
DATA_START_ROW = 2      # 数据起始 Excel 行


# ───────────────────────── 链 ① 第一步:清洗 + 捕获血缘 ─────────────────────────
def load_canonical(path: str, sheet: str) -> pd.DataFrame:
    """读 sheet → 轻清洗 → 返回 canonical DataFrame,带 __row_id / __excel_row。

    __row_id   稳定身份(随行走,排序不变)—— diff 的 key
    __excel_row 该行在原件里的物理 Excel 行号 —— 导出回写定位用
    """
    df = pd.read_excel(path, sheet_name=sheet, header=0, dtype=object)
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

    # 删合计/小计行(模拟 excel_ingest._drop_total_rows):任一格精确等于关键词就删。
    # 这些行被剔出编辑基线 → 不可编辑;但导出时仍原样留在原件里(决策 6)。
    totals = {"小计", "合计", "总计", "汇总", "total", "subtotal", "sum"}
    df = df[~df.apply(
        lambda r: any(isinstance(v, str) and v.strip().lower() in totals for v in r),
        axis=1,
    )]

    # reset_index 前,原始 0-based 位置 → excel_row = 位置 + DATA_START_ROW
    excel_rows = [int(p) + DATA_START_ROW for p in df.index]
    df = df.reset_index(drop=True)

    # 迷你类型清洗:去千分位/%/空格,≥80% 能转数值 → 整列转数值(模拟 _coerce_column)
    for c in df.columns:
        s = df[c].dropna().astype(str)
        if not len(s):
            continue
        cleaned = s.str.replace(r"[,\s]", "", regex=True).str.replace("%", "", regex=False)
        if pd.to_numeric(cleaned, errors="coerce").notna().mean() >= 0.8:
            full = (df[c].astype(str)
                    .str.replace(r"[,\s]", "", regex=True)
                    .str.replace("%", "", regex=False))
            df[c] = pd.to_numeric(full, errors="coerce")

    df.insert(0, "__excel_row", excel_rows)
    df.insert(0, "__row_id", [f"r{i + 1}" for i in range(len(df))])
    return df


def col_to_excel_col(path: str, sheet: str) -> dict[str, int]:
    """从原件表头行读 列名 → Excel 列号(1-based)。真实实现也是这套匹配。"""
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    mapping: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(HEADER_ROW, c).value
        if name is not None:
            mapping[str(name)] = c
    wb.close()
    return mapping


# ───────────────────────── 链 ① 第二步:DuckDB 改数据 ─────────────────────────
def apply_edits(df0: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """把 canonical 物化进 DuckDB,跑三种编辑,返回最终态 df1 + 操作说明。"""
    con = duckdb.connect(":memory:")
    con.register("src", df0)
    con.execute('CREATE TABLE t AS SELECT * FROM src')

    ops = []
    # UPDATE:华东工厂 A 型 产量 → 14000
    con.execute('UPDATE t SET "产量" = 14000 '
                'WHERE "工厂" = \'华东工厂\' AND "产品" = \'A 型\'')
    ops.append("UPDATE 产量=14000 @华东工厂/A型")

    # DELETE:删掉 华南工厂 C 型 这一行
    con.execute('DELETE FROM t WHERE "工厂" = \'华南工厂\' AND "产品" = \'C 型\'')
    ops.append("DELETE 华南工厂/C型")

    # ALTER ADD COLUMN:新增「达标」,合格率>=98 → 是,否则 否
    con.execute('ALTER TABLE t ADD COLUMN "达标" VARCHAR')
    con.execute('UPDATE t SET "达标" = CASE WHEN "合格率" >= 98 THEN \'是\' ELSE \'否\' END')
    ops.append("ALTER 加列 达标(合格率>=98?是:否)")

    df1 = con.execute('SELECT * FROM t ORDER BY __excel_row NULLS LAST').fetch_df()
    con.close()
    return df1, ops


# ───────────────────────── 链 ① 第三步:单元格级 diff ─────────────────────────
def compute_diff(df0: pd.DataFrame, df1: pd.DataFrame) -> dict:
    """对比最终态 vs 原始态,产出 cell-level diff(按 __row_id)。"""
    meta = {"__row_id", "__excel_row"}
    cols0 = [c for c in df0.columns if c not in meta]
    cols1 = [c for c in df1.columns if c not in meta]

    added_cols = [c for c in cols1 if c not in cols0]
    dropped_cols = [c for c in cols0 if c not in cols1]

    by_id0 = {r["__row_id"]: r for _, r in df0.iterrows()}
    by_id1 = {r["__row_id"]: r for _, r in df1.iterrows()}

    deleted = [by_id0[rid]["__excel_row"] for rid in by_id0 if rid not in by_id1]
    new_rows = [rid for rid in by_id1 if rid not in by_id0]

    changes = []  # {excel_row, col, old, new}
    for rid, r1 in by_id1.items():
        if rid not in by_id0:
            continue
        r0 = by_id0[rid]
        er = r0["__excel_row"]
        for c in cols0:
            if c in cols1 and not _eq(r0[c], r1[c]):
                changes.append({"excel_row": int(er), "col": c,
                                "old": r0[c], "new": r1[c]})
    return {
        "added_cols": added_cols, "dropped_cols": dropped_cols,
        "deleted_excel_rows": [int(x) for x in deleted],
        "new_row_ids": new_rows, "cell_changes": changes,
    }


def _eq(a, b) -> bool:
    if pd.isna(a) and pd.isna(b):
        return True
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return str(a) == str(b)


# ───────────────────────── §6.1 镜像原列表示 ─────────────────────────
def _render_like_column(ws, col_idx: int, skip_row: int, value):
    """看同列某个未改单元格的原始形态,把 value 渲染成"长得一样"的写法。

    原列是文本带千分位 → 返回 '14,000' 文本;带 % → '88%' 文本;
    原列是真数字 → 返回数字本身(沿用原 number_format,Excel 右对齐)。
    """
    sample = None
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if r == skip_row:
            continue
        v = ws.cell(r, col_idx).value
        if v is not None:
            sample = v
            break

    if isinstance(sample, str):
        has_comma = "," in sample
        has_pct = sample.strip().endswith("%")
        if isinstance(value, (int, float)):
            if has_pct:
                num = value if value % 1 else int(value)
                return f"{num}%"
            if has_comma:
                return f"{int(value):,}" if float(value).is_integer() else f"{value:,}"
        return str(value)
    # 原列是真数字(或空)→ 写数字,保留该列原 number_format(不动)
    return value


# ───────────────────────── 链 ② openpyxl 保样式回写 ─────────────────────────
def patch_workbook(in_path: str, out_path: str, sheet: str,
                   df1: pd.DataFrame, diff: dict, colmap: dict[str, int]) -> None:
    """打开原件 → 按 diff 逐格回写(保样式)→ 另存 out_path(原件不动)。

    顺序很重要:① 改值 → ② 加列(含全行) → ③ 删行(倒序),避免行号错位。
    """
    wb = load_workbook(in_path)
    ws = wb[sheet]

    # ① 改值:定位单元格改 .value,样式天然保留。
    # §6.1:镜像原列表示 —— 原列是文本(13,478 / 98.5%)就把新值也写成同形态文本串,
    # 保持左对齐+绿三角与邻居一致;原列是真数字就写数字、沿用原 number_format。
    for ch in diff["cell_changes"]:
        c = colmap.get(ch["col"])
        if c is None:
            continue
        cell = ws.cell(ch["excel_row"], c)
        cell.value = _render_like_column(ws, c, ch["excel_row"], ch["new"])

    # ② 加列:末列右侧新增,表头 + 每行值,样式克隆左邻列
    for col_name in diff["added_cols"]:
        new_idx = ws.max_column + 1
        left = new_idx - 1
        h = ws.cell(HEADER_ROW, new_idx)
        h.value = col_name
        h._style = copy(ws.cell(HEADER_ROW, left)._style)
        for _, row in df1.iterrows():
            er = row["__excel_row"]
            if pd.isna(er):           # 新插入行无 excel_row,spike 不涉及
                continue
            er = int(er)
            t = ws.cell(er, new_idx)
            t.value = None if pd.isna(row[col_name]) else row[col_name]
            t._style = copy(ws.cell(er, left)._style)

    # ③ 删行:倒序删,避免删前面导致后面行号错位
    for er in sorted(diff["deleted_excel_rows"], reverse=True):
        ws.delete_rows(er)

    # 留存的公式让 Excel 打开时重算(spike 无公式,设上无害)
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)
    wb.close()


# ───────────────────────── 验证 ─────────────────────────
def verify(out_path: str, sheet: str) -> None:
    """读回结果,打印关键单元格,验证「没改的保持原文本、改过的是新值」。"""
    wb = load_workbook(out_path)
    ws = wb[sheet]
    print("\n========== 验证:读回 result.xlsx ==========")
    print(f"sheet「{sheet}」尺寸:{ws.dimensions}  最大列={ws.max_column}")
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        fmts = [ws.cell(r, c).number_format for c in range(1, ws.max_column + 1)]
        print(f"  行{r}: {vals}")
        if r == 2:
            print(f"        (数字格式: {fmts})")
    wb.close()


def main() -> None:
    in_path = sys.argv[1] if len(sys.argv) > 1 else "../../normal_production.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "spike_result.xlsx"

    print(f"输入原件: {in_path}\n输出结果: {out_path}\n")

    df0 = load_canonical(in_path, SHEET)
    colmap = col_to_excel_col(in_path, SHEET)
    print("===== 链①-1 canonical(清洗+血缘) =====")
    print(df0.to_string())
    print(f"\n列→Excel列号映射: {colmap}")

    df1, ops = apply_edits(df0)
    print("\n===== 链①-2 DuckDB 编辑后(最终态) =====")
    for o in ops:
        print(f"  · {o}")
    print(df1.to_string())

    diff = compute_diff(df0, df1)
    print("\n===== 链①-3 单元格级 diff =====")
    print(f"  加列: {diff['added_cols']}   删列: {diff['dropped_cols']}")
    print(f"  删行(excel_row): {diff['deleted_excel_rows']}")
    print(f"  新行: {diff['new_row_ids']}")
    print(f"  改格 {len(diff['cell_changes'])} 处:")
    for ch in diff["cell_changes"]:
        print(f"    行{ch['excel_row']} 列「{ch['col']}」: {ch['old']!r} → {ch['new']!r}")

    patch_workbook(in_path, out_path, SHEET, df1, diff, colmap)
    print("\n===== 链② openpyxl 已保样式回写并存盘 =====")

    verify(out_path, SHEET)
    print("\n✅ spike 跑通。关注点:① 未改的格(如其它行产量)是否仍是原始文本 "
          "'9,061';② 改过的 14000 是否带千分位;③ 删行后行数 -1;④ 末尾多出「达标」列。")


if __name__ == "__main__":
    main()
