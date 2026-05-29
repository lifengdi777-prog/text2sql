"""生成两个测试 Excel 用于验证上传/清洗/入库链路。

运行:uv run python -m test.make_test_excels
输出:
  test_data/normal_production.xlsx   正常表(表头第 0 行,2 个 sheet,含千分位/百分号/合计行)
  test_data/sales_with_dates.xlsx    带日期 + 高基数列(测时间维度 + top_K)
"""
from pathlib import Path

from openpyxl import Workbook

OUT_DIR = Path("test_data")


def make_normal_production() -> Path:
    """正常生产数据:2 个 sheet,含常见脏数据特征。"""
    wb = Workbook()

    # Sheet 1:生产明细 —— 千分位/百分号/合计行
    ws1 = wb.active
    ws1.title = "生产明细"
    ws1.append(["工厂", "产品", "产量", "合格率", "不良率"])
    rows1 = [
        ("华东工厂", "A 型", "13,478", "98.5%", "1.5%"),
        ("华东工厂", "B 型", "9,061",  "97.2%", "2.8%"),
        ("华北工厂", "A 型", "7,452",  "96.8%", "3.2%"),
        ("华北工厂", "B 型", "5,820",  "95.4%", "4.6%"),
        ("华南工厂", "A 型", "4,231",  "98.1%", "1.9%"),
        ("华南工厂", "C 型", "3,560",  "94.5%", "5.5%"),
        ("西南工厂", "A 型", "2,588",  "97.8%", "2.2%"),
        ("西南工厂", "C 型", "1,890",  "93.7%", "6.3%"),
        ("合计",    "",     "48,080", "",      ""),     # ← 应被去掉
    ]
    for r in rows1:
        ws1.append(r)

    # Sheet 2:设备台账 —— 字符串/日期混合
    ws2 = wb.create_sheet("设备台账")
    ws2.append(["设备编号", "设备类型", "状态", "启用日期"])
    rows2 = [
        ("E001", "注塑机",   "运行", "2024-01-15"),
        ("E002", "注塑机",   "维护", "2024-03-20"),
        ("E003", "数控机床", "运行", "2025-06-01"),
        ("E004", "检测设备", "停机", "2023-11-10"),
        ("E005", "数控机床", "运行", "2025-09-12"),
    ]
    for r in rows2:
        ws2.append(r)

    path = OUT_DIR / "normal_production.xlsx"
    wb.save(path)
    return path


def make_sales_with_dates() -> Path:
    """销售流水:测时间列(temporal) + 高基数产品列(top_K)+ 数值范围。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"
    ws.append(["日期", "工厂", "产品名称", "销量", "单价"])

    import random
    random.seed(42)
    factories = ["华东工厂", "华北工厂", "华南工厂", "西南工厂"]
    products = [f"Product-{i:03d}" for i in range(1, 81)]   # 80 个产品(>50,触发 top_K)
    days = ["2025-01-15", "2025-02-20", "2025-03-10", "2025-04-05", "2025-05-22"]

    for _ in range(120):
        ws.append((
            random.choice(days),
            random.choice(factories),
            random.choice(products),
            random.randint(10, 500),
            round(random.uniform(99, 9999), 2),
        ))

    path = OUT_DIR / "sales_with_dates.xlsx"
    wb.save(path)
    return path


def make_ecommerce_orders() -> Path:
    """电商订单:一个数据集尽量覆盖多种问题类型(给问数功能做「综合测试集」)。

    订单明细(300 行):
      - 时间维度:下单日期(temporal,测 between / min / max)+ 月份(测按月趋势)
      - 多维分组:地区 / 品类 / 支付方式 / 是否会员 / 订单状态(低基数 categorical)
      - 高基数:商品名称(测 icontains / top_K)、客户ID(测 nunique 去重客户数)
      - 数值:数量 / 单价 / 金额(测 sum / mean / max / min / median / between / gt)
      - 空值:约 5% 支付方式留空(测 is_null / not_null)
    客户信息(60 行):
      - 年龄(测区间 between)、会员等级(有序 categorical)、性别、城市
    注意:两个 sheet 都有「客户ID」,但 ComputeSpec 不做 join —— 每个问题只在单 sheet 内回答。
    """
    import random
    from datetime import date, timedelta

    random.seed(7)
    wb = Workbook()

    # ── Sheet 1:订单明细 ──
    ws = wb.active
    ws.title = "订单明细"
    ws.append(["订单号", "下单日期", "月份", "客户ID", "地区", "品类",
               "商品名称", "数量", "单价", "金额", "支付方式", "订单状态", "是否会员"])

    regions = ["华东", "华北", "华南", "华中", "西南", "东北"]
    categories = {
        "数码": ["手机", "蓝牙耳机", "平板电脑", "充电器", "智能手表"],
        "服饰": ["纯棉T恤", "牛仔裤", "连帽卫衣", "运动鞋", "羽绒外套"],
        "家居": ["护眼台灯", "抱枕", "收纳盒", "陶瓷餐具", "地毯"],
        "食品": ["每日坚果", "曲奇饼干", "挂耳咖啡", "龙井茶叶", "黑巧克力"],
        "美妆": ["哑光口红", "补水面膜", "淡香水", "轻薄粉底", "大地色眼影"],
    }
    pays = ["支付宝", "微信", "银联", "货到付款"]
    statuses = ["已完成", "已完成", "已完成", "配送中", "已退款", "已取消"]  # 加权:已完成偏多
    members = ["是", "否"]
    customer_ids = [f"C{1000 + i}" for i in range(60)]   # 60 个客户 → 测 nunique
    start = date(2025, 1, 1)

    for i in range(300):
        cat = random.choice(list(categories.keys()))
        qty = random.randint(1, 10)
        price = round(random.uniform(9.9, 1999), 2)
        d = start + timedelta(days=random.randint(0, 364))     # 散布全年
        pay = random.choice(pays)
        if random.random() < 0.05:                             # ~5% 留空,测 is_null
            pay = None
        ws.append([
            f"ORD{20250000 + i}",
            d.isoformat(),
            f"{d.year}-{d.month:02d}",                         # 月份:零填充,字符串排序正确
            random.choice(customer_ids),
            random.choice(regions),
            cat,
            random.choice(categories[cat]),
            qty,
            price,
            round(qty * price, 2),
            pay,
            random.choice(statuses),
            random.choice(members),
        ])

    # ── Sheet 2:客户信息 ──
    ws2 = wb.create_sheet("客户信息")
    ws2.append(["客户ID", "客户姓名", "注册日期", "城市", "年龄", "性别", "会员等级"])
    cities = ["北京", "上海", "广州", "深圳", "杭州", "成都", "武汉", "西安"]
    levels = ["普通", "银卡", "金卡", "铂金"]
    genders = ["男", "女"]
    for idx, cid in enumerate(customer_ids):
        reg_d = start + timedelta(days=random.randint(-700, -1))
        ws2.append([
            cid,
            f"客户{idx + 1:02d}",
            reg_d.isoformat(),
            random.choice(cities),
            random.randint(18, 65),
            random.choice(genders),
            random.choice(levels),
        ])

    path = OUT_DIR / "ecommerce_orders.xlsx"
    wb.save(path)
    return path


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    p1 = make_normal_production()
    p2 = make_sales_with_dates()
    p3 = make_ecommerce_orders()
    print(f"[OK] {p1}")
    print(f"     2 sheet(生产明细 / 设备台账),含千分位/百分号/合计行")
    print(f"[OK] {p2}")
    print(f"     1 sheet(销售明细 120 行),含 80 个产品(触发 top_K)+ 日期列(temporal)")
    print(f"[OK] {p3}")
    print(f"     2 sheet(订单明细 300 行 / 客户信息 60 行),覆盖时间趋势/多维分组/范围过滤/去重/空值")


if __name__ == "__main__":
    main()
